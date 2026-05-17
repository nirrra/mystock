from __future__ import annotations

from datetime import date
from pathlib import Path
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook

from stocks_analyzer.sector_tracking_workbook import (
    BUY_SCORE_SHEET,
    LONG_MAINLINE_SHEET,
    SHORT_MAINLINE_SHEET,
    sector_daily_tracking_workbook_path,
    sector_intraday_tracking_workbook_path,
    write_sector_daily_tracking_workbook,
    write_sector_intraday_tracking_workbook,
)


ROOT = Path(__file__).resolve().parents[1]


def _make_workspace_tmp_dir(name: str) -> Path:
    path = ROOT / ".tmp_tests" / f"{name}_{uuid4().hex}"
    path.mkdir(parents=True, exist_ok=True)
    return path


def test_daily_sector_tracking_workbook_writes_three_tracking_sheets() -> None:
    tmp_path = _make_workspace_tmp_dir("sector_tracking_daily")
    payload = _sector_payload(
        trade_date=date(2026, 5, 15),
        sectors=[
            _sector_item("industry", "商业航天", "881001", long_score=96.0, short_score=72.0, p9=81.0, short=True),
            _sector_item("concept", "机器人", "885001", long_score=91.0, short_score=48.0, p9=66.0),
        ],
    )

    target = write_sector_daily_tracking_workbook(
        project_root=tmp_path,
        trade_date=date(2026, 5, 15),
        sector_payload=payload,
    )

    assert target == sector_daily_tracking_workbook_path(tmp_path)
    workbook = load_workbook(target, read_only=True, data_only=True)
    assert workbook.sheetnames == [LONG_MAINLINE_SHEET, SHORT_MAINLINE_SHEET, BUY_SCORE_SHEET]
    workbook.close()

    long_frame = pd.read_excel(target, sheet_name=LONG_MAINLINE_SHEET, dtype={"板块代码": str, "龙头编号": str})
    short_frame = pd.read_excel(target, sheet_name=SHORT_MAINLINE_SHEET, dtype={"板块代码": str, "龙头编号": str})
    buy_frame = pd.read_excel(target, sheet_name=BUY_SCORE_SHEET, dtype={"板块代码": str, "龙头编号": str})

    assert long_frame["板块名称"].tolist() == ["商业航天", "机器人"]
    assert short_frame["板块名称"].tolist() == ["商业航天", "机器人"]
    assert buy_frame["板块名称"].tolist() == ["商业航天", "机器人"]
    assert long_frame.loc[0, "入选原因"] == "长期主线/短期主线/P9高买入分"
    assert long_frame.loc[0, "龙头编号"] == "600001/600002/600003"


def test_daily_sector_tracking_workbook_replaces_same_date_rows() -> None:
    tmp_path = _make_workspace_tmp_dir("sector_tracking_replace")
    first_payload = _sector_payload(
        trade_date=date(2026, 5, 15),
        sectors=[_sector_item("industry", "商业航天", "881001", long_score=96.0, short_score=72.0, p9=81.0, short=True)],
    )
    second_payload = _sector_payload(
        trade_date=date(2026, 5, 15),
        sectors=[_sector_item("concept", "机器人", "885001", long_score=93.0, short_score=55.0, p9=77.0, short=True)],
    )

    write_sector_daily_tracking_workbook(
        project_root=tmp_path,
        trade_date=date(2026, 5, 15),
        sector_payload=first_payload,
    )
    write_sector_daily_tracking_workbook(
        project_root=tmp_path,
        trade_date=date(2026, 5, 15),
        sector_payload=second_payload,
    )

    frame = pd.read_excel(sector_daily_tracking_workbook_path(tmp_path), sheet_name=LONG_MAINLINE_SHEET)
    assert frame["板块名称"].tolist() == ["机器人"]


def test_intraday_sector_tracking_workbook_merges_leader_strength() -> None:
    tmp_path = _make_workspace_tmp_dir("sector_tracking_intraday")
    payload = _sector_payload(
        trade_date=date(2026, 5, 15),
        sectors=[_sector_item("industry", "商业航天", "881001", long_score=96.0, short_score=72.0, p9=81.0, short=True)],
    )
    intraday = pd.DataFrame(
        [
            {
                "日期": "2026-05-15",
                "板块类型": "行业",
                "板块名称": "商业航天",
                "龙头盘中平均涨幅%": 3.21,
                "有效龙头数": 3,
            }
        ]
    )

    target = write_sector_intraday_tracking_workbook(
        project_root=tmp_path,
        trade_date=date(2026, 5, 15),
        sector_payload=payload,
        intraday_strength=intraday,
    )

    assert target == sector_intraday_tracking_workbook_path(tmp_path)
    frame = pd.read_excel(target, sheet_name=LONG_MAINLINE_SHEET, dtype={"板块代码": str, "龙头编号": str})
    assert frame.loc[0, "板块名称"] == "商业航天"
    assert float(frame.loc[0, "龙头盘中平均涨幅%"]) == 3.21
    assert int(frame.loc[0, "有效龙头数"]) == 3


def _sector_payload(*, trade_date: date, sectors: list[dict[str, object]]) -> dict[str, object]:
    return {
        "trade_date": trade_date.isoformat(),
        "sector_count": len(sectors),
        "sectors": sectors,
    }


def _sector_item(
    sector_type: str,
    sector_name: str,
    sector_label: str,
    *,
    long_score: float,
    short_score: float,
    p9: float,
    short: bool = False,
) -> dict[str, object]:
    return {
        "trade_date": "2026-05-15",
        "sector_type": sector_type,
        "sector_name": sector_name,
        "sector_label": sector_label,
        "sector_avg_pct_change": 1.23,
        "sector_amount_weighted_pct_change": 1.68,
        "sector_up_ratio": 0.62,
        "long_mainline_score_100": long_score,
        "short_mainline_score_100": short_score,
        "phase9_score_100": p9,
        "phase9_rank": 5,
        "selected_as_long_mainline": True,
        "selected_as_short_mainline": short,
        "selected_as_phase9_buy": p9 >= 70,
        "leader_symbols": ["600001", "600002", "600003"],
        "leader_names": ["龙一", "龙二", "龙三"],
        "member_count": 20,
        "valid_count": 18,
    }
