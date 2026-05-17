from __future__ import annotations

from datetime import date
from pathlib import Path
from uuid import uuid4

import pandas as pd
from openpyxl import Workbook, load_workbook

from stocks_analyzer.daily_returns import full_market_daily_returns_path
from stocks_analyzer.track_stock import update_track_stock_workbook


ROOT = Path(__file__).resolve().parents[1]


def _make_workspace_tmp_dir(name: str) -> Path:
    path = ROOT / ".tmp_tests" / f"{name}_{uuid4().hex}"
    path.mkdir(parents=True, exist_ok=True)
    return path


def test_track_stock_daily_sheet_backfills_daily_return_from_local_bars() -> None:
    project_root = _make_workspace_tmp_dir("track_stock_daily_return")
    daily_dir = project_root / "data" / "daily"
    daily_dir.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(
        [
            {"trade_date": "2026-05-14", "symbol": "600000", "close": 10.0, "volume": 1000.0, "amount": 10000.0},
            {"trade_date": "2026-05-15", "symbol": "600000", "close": 10.5, "volume": 2000.0, "amount": 21000.0},
        ]
    ).to_parquet(daily_dir / "600000.parquet", index=False)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "股票代码"
    sheet["A2"] = "600000"
    workbook.save(project_root / "track_stock.xlsx")

    result = update_track_stock_workbook(project_root=project_root, trade_date=date(2026, 5, 15), mode="daily")

    assert result.output_rows == 1
    assert full_market_daily_returns_path(project_root, date(2026, 5, 15)).exists()
    updated = load_workbook(project_root / "track_stock.xlsx")
    sheet2 = updated["Sheet2"]
    headers = [cell.value for cell in sheet2[1]]
    row = {header: sheet2.cell(row=2, column=index).value for index, header in enumerate(headers, start=1)}
    assert row["股票代码"] == "600000"
    assert row["当日涨幅%"] == 5
    assert row["收盘价"] == 10.5
