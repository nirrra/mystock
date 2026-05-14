from __future__ import annotations

import json
import math
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .atr import build_atr_snapshot_row
from .full_market_limit_up_3d import limit_up_3d_model_path, predict_limit_up_3d_opportunity
from .full_market_return import predict_alpha158_qlib_return
from .full_market_risk import predict_barrier_risk, predict_tail_risk
from .indicators import add_indicators
from .intraday_update import INTRADAY_DATA_INTERFACES, run_intraday_update
from .macd_divergence import summarize_recent_macd_divergence
from .phase4_rolling import PHASE4_ROLLING_COLUMNS, PHASE4_ROLLING_RANK_COLUMNS, merge_phase4_rolling_frame
from .phase_display import normalize_symbol
from .position_sizing import RECOMMENDED_POSITION_PERCENT_FIELD, add_recommended_position_percent
from .storage import DailyBarsReadError, Storage
from .track_stock import DEFAULT_TRACK_STOCK_FILENAME, TRACK_INPUT_SHEET
from .watchlist import (
    _prepare_atr_frame,
    _prepare_macd_frame,
    _prepare_phase4_predictions,
    _prepare_phase8_predictions,
    _prepare_phase_risk_predictions,
    add_centered_risk_scores,
    build_intraday_pool_candidates,
    find_latest_intraday_pool_before,
    intraday_pool_path,
    write_intraday_pool,
)

DEFAULT_INTRADAY_REPORT_KEEP_DATES = 10
_DATED_REPORT_PATTERNS = (
    re.compile(r"^intraday_pool_screening_(\d{4}-\d{2}-\d{2})\.csv$"),
    re.compile(r"^intraday_track_stock_(\d{4}-\d{2}-\d{2})\.csv$"),
)


@dataclass(slots=True)
class IntradayScreeningResult:
    trade_date: date
    source_pool_path: Path
    output_path: Path
    track_stock_path: Path | None
    candidate_count: int
    pool_candidate_count: int
    track_stock_count: int
    intraday_updated_count: int
    missing_intraday_symbols: list[str]
    cleaned_report_files: int
    phase1_path: Path
    phase2_path: Path
    phase4_path: Path
    phase8_path: Path | None
    full_market_pool_refreshed: bool = False
    full_market_scanned_count: int = 0


@dataclass(slots=True)
class _AnalysisResult:
    frame: pd.DataFrame
    candidates: list[dict[str, object]]
    intraday: pd.DataFrame
    phase1: pd.DataFrame
    phase2: pd.DataFrame
    phase4: pd.DataFrame
    phase8: pd.DataFrame
    macd: pd.DataFrame
    atr: pd.DataFrame
    phase1_path: Path
    phase2_path: Path
    phase4_path: Path
    phase8_path: Path | None


def run_intraday_screening(
    *,
    storage: Storage,
    project_root: Path,
    trade_date: date,
    data_interface: str = "sina_raw",
    limit: int | None = None,
    skip_intraday_update: bool = False,
    timeout_seconds: float = 15.0,
    chunk_size: int = 50,
    output: Path | None = None,
    report_keep_dates: int | None = DEFAULT_INTRADAY_REPORT_KEEP_DATES,
    refresh_full_market_pool: bool = False,
) -> IntradayScreeningResult:
    updated_count = 0
    missing_update_symbols: list[str] = []
    full_market_scanned_count = 0

    if refresh_full_market_pool:
        (
            source_pool_path,
            pool_payload,
            updated_count,
            missing_update_symbols,
            full_market_scanned_count,
        ) = _refresh_full_market_intraday_pool(
            storage=storage,
            project_root=project_root,
            trade_date=trade_date,
            data_interface=data_interface,
            limit=limit,
            skip_intraday_update=skip_intraday_update,
            timeout_seconds=timeout_seconds,
            chunk_size=chunk_size,
        )
    else:
        source_pool_path, pool_payload = _load_intraday_pool_for_screening(project_root, trade_date)

    pool_candidates = _previous_pool_candidates(pool_payload, limit=None)
    candidates = pool_candidates
    if limit is not None:
        candidates = candidates[: max(int(limit), 0)]
    tracked_symbols = _load_tracked_symbols(project_root)
    candidates = _merge_tracked_candidates(storage=storage, candidates=candidates, tracked_symbols=tracked_symbols)
    symbols = [candidate["symbol"] for candidate in candidates]
    if not symbols:
        raise RuntimeError(f"No previous intraday-pool symbols found in {source_pool_path}")

    if not skip_intraday_update and not refresh_full_market_pool:
        update_result = run_intraday_update(
            storage=storage,
            project_root=project_root,
            source=data_interface,
            symbols=symbols,
            timeout_seconds=timeout_seconds,
            chunk_size=chunk_size,
        )
        updated_count = len(update_result.updated_symbols)
        missing_update_symbols = update_result.failed_symbols

    output_dir = project_root / "reports" / "intraday_screening"
    output_dir.mkdir(parents=True, exist_ok=True)
    cache_dir = storage.paths.intraday_dir / "screening_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)

    analysis = _run_candidate_analysis(
        storage=storage,
        project_root=project_root,
        trade_date=trade_date,
        candidates=candidates,
        output_dir=cache_dir,
        file_tag="pool",
    )
    output_path = output if output is not None else output_dir / f"intraday_pool_screening_{trade_date.isoformat()}.csv"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    analysis.frame.to_csv(output_path, index=False, encoding="utf-8-sig")
    track_stock_path = _save_intraday_track_stock(output_dir=output_dir, trade_date=trade_date, frame=analysis.frame)
    cleaned_report_files = _cleanup_intraday_screening_reports(
        output_dir,
        keep_dates=report_keep_dates,
        preserve_paths=(output_path, track_stock_path),
    )

    intraday = _load_intraday_snapshot_frame(storage, symbols=symbols)
    missing_intraday = [symbol for symbol in symbols if symbol not in set(intraday.get("symbol", pd.Series(dtype=str)).astype(str))]
    return IntradayScreeningResult(
        trade_date=trade_date,
        source_pool_path=source_pool_path,
        output_path=output_path,
        track_stock_path=track_stock_path,
        candidate_count=len(analysis.frame),
        pool_candidate_count=len(pool_candidates),
        track_stock_count=len(tracked_symbols),
        intraday_updated_count=updated_count,
        missing_intraday_symbols=sorted(set(missing_update_symbols + missing_intraday)),
        cleaned_report_files=cleaned_report_files,
        phase1_path=analysis.phase1_path,
        phase2_path=analysis.phase2_path,
        phase4_path=analysis.phase4_path,
        phase8_path=analysis.phase8_path,
        full_market_pool_refreshed=refresh_full_market_pool,
        full_market_scanned_count=full_market_scanned_count,
    )


class _IntradayOverlayStorage:
    def __init__(self, *, storage: Storage, universe: pd.DataFrame, trade_date: date) -> None:
        self._storage = storage
        self._universe = universe.copy()
        self._trade_date = trade_date
        self.paths = storage.paths

    def load_universe(self) -> pd.DataFrame:
        return self._universe.copy()

    def load_daily_bars(self, symbol: str) -> pd.DataFrame:
        normalized_symbol = normalize_symbol(symbol)
        daily = self._storage.load_daily_bars(normalized_symbol).copy()
        daily["trade_date"] = pd.to_datetime(daily["trade_date"], errors="coerce")
        daily = daily.dropna(subset=["trade_date"])
        daily = daily[daily["trade_date"].dt.date < self._trade_date].copy()

        try:
            intraday = self._storage.load_intraday_bars(normalized_symbol)
        except FileNotFoundError:
            return daily.sort_values("trade_date").reset_index(drop=True)

        provisional = _normalize_intraday_bar_for_daily(intraday, symbol=normalized_symbol, trade_date=self._trade_date)
        if provisional.empty:
            return daily.sort_values("trade_date").reset_index(drop=True)

        merged = pd.concat([daily, provisional], ignore_index=True)
        merged["trade_date"] = pd.to_datetime(merged["trade_date"], errors="coerce")
        merged = merged.dropna(subset=["trade_date"]).drop_duplicates("trade_date", keep="last")
        return merged.sort_values("trade_date").reset_index(drop=True)


def _load_previous_intraday_pool(project_root: Path, trade_date: date) -> tuple[Path, dict[str, object]]:
    _, path = find_latest_intraday_pool_before(project_root=project_root, trade_date=trade_date)
    return path, json.loads(path.read_text(encoding="utf-8"))


def _load_intraday_pool_for_screening(project_root: Path, trade_date: date) -> tuple[Path, dict[str, object]]:
    today_path = intraday_pool_path(project_root, trade_date)
    if today_path.exists():
        payload = json.loads(today_path.read_text(encoding="utf-8"))
        if _is_intraday_full_market_pool(payload):
            return today_path, payload
    return _load_previous_intraday_pool(project_root, trade_date)


def _is_intraday_full_market_pool(payload: dict[str, object]) -> bool:
    policy = payload.get("selection_policy") if isinstance(payload, dict) else {}
    if not isinstance(policy, dict):
        return False
    return str(policy.get("source_scope", "")).strip() == "intraday_full_market"


def _refresh_full_market_intraday_pool(
    *,
    storage: Storage,
    project_root: Path,
    trade_date: date,
    data_interface: str,
    limit: int | None,
    skip_intraday_update: bool,
    timeout_seconds: float,
    chunk_size: int,
) -> tuple[Path, dict[str, object], int, list[str], int]:
    full_market_candidates = _full_market_candidates(storage, limit=limit)
    symbols = [candidate["symbol"] for candidate in full_market_candidates]
    if not symbols:
        raise RuntimeError("No full-market symbols found in local universe.")

    updated_count = 0
    missing_update_symbols: list[str] = []
    if not skip_intraday_update:
        update_result = run_intraday_update(
            storage=storage,
            project_root=project_root,
            source=data_interface,
            symbols=symbols,
            timeout_seconds=timeout_seconds,
            chunk_size=chunk_size,
        )
        updated_count = len(update_result.updated_symbols)
        missing_update_symbols = update_result.failed_symbols

    cache_dir = storage.paths.intraday_dir / "screening_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    analysis = _run_candidate_analysis(
        storage=storage,
        project_root=project_root,
        trade_date=trade_date,
        candidates=full_market_candidates,
        output_dir=cache_dir,
        file_tag="full_market",
    )
    phase5_path = project_root / "reports" / "full_market_model" / "mcd_crash_annual_measures.csv"
    pattern_path, pattern_frame = _load_latest_pattern_frame(project_root=project_root, trade_date=trade_date)
    payload = build_intraday_pool_candidates(
        trade_date=trade_date,
        pattern_frame=pattern_frame,
        phase1_predictions=analysis.phase1,
        phase2_predictions=analysis.phase2,
        phase4_predictions=analysis.phase4,
        phase8_predictions=analysis.phase8,
        phase7_prediction=pd.DataFrame(),
        phase5_measures=_read_optional_csv(phase5_path),
        macd_frame=analysis.macd,
        atr_frame=analysis.atr,
        source_files={
            "pattern": str(pattern_path) if pattern_path is not None else "",
            "phase1": str(analysis.phase1_path),
            "phase2": str(analysis.phase2_path),
            "phase4": str(analysis.phase4_path),
            "phase8": str(analysis.phase8_path) if analysis.phase8_path is not None else "",
            "phase5": str(phase5_path),
            "macd": "intraday_overlay",
            "atr": "intraday_overlay",
        },
    )
    selection_policy = dict(payload.get("selection_policy", {}))
    selection_policy["source_scope"] = "intraday_full_market"
    selection_policy["full_market_scan_symbols"] = int(len(full_market_candidates))
    selection_policy["full_market_pool_date"] = trade_date.isoformat()
    payload["selection_policy"] = selection_policy
    filter_summary = dict(payload.get("filter_summary", {}))
    filter_summary["full_market_scan_symbols"] = int(len(full_market_candidates))
    filter_summary["full_market_intraday_updated"] = int(updated_count)
    filter_summary["full_market_intraday_missing"] = int(len(missing_update_symbols))
    payload["filter_summary"] = filter_summary

    target = write_intraday_pool(project_root=project_root, trade_date=trade_date, picker_payload=payload)
    written_payload = json.loads(target.read_text(encoding="utf-8"))
    return target, written_payload, updated_count, missing_update_symbols, len(full_market_candidates)


def _full_market_candidates(storage: Storage, *, limit: int | None) -> list[dict[str, object]]:
    universe = storage.load_universe().copy()
    if "symbol" not in universe.columns:
        return []
    universe["symbol"] = universe["symbol"].map(normalize_symbol)
    universe = universe[universe["symbol"].astype(str).str.len().eq(6)].drop_duplicates("symbol", keep="first")
    if "name" not in universe.columns:
        universe["name"] = ""
    if limit is not None:
        universe = universe.head(max(int(limit), 0))
    return [
        {
            "symbol": str(row.get("symbol", "")),
            "name": str(row.get("name", "") or ""),
            "universe_rank": index,
            "prev_source": "full_market_scan",
            "prev_source_tags": "",
            "prev_pattern_match": False,
            "prev_pattern_id": "",
            "prev_pattern_ids": "",
            "prev_patterns": "",
            "prev_reason": "",
            "track_stock": False,
        }
        for index, row in enumerate(universe.to_dict("records"), start=1)
    ]


def _load_latest_pattern_frame(*, project_root: Path, trade_date: date) -> tuple[Path | None, pd.DataFrame]:
    pattern_dir = project_root / "reports" / "patterns"
    dated: list[tuple[date, Path]] = []
    for path in pattern_dir.glob("patterns_all_*.csv"):
        match = re.fullmatch(r"patterns_all_(\d{4}-\d{2}-\d{2})\.csv", path.name)
        if not match:
            continue
        try:
            parsed = date.fromisoformat(match.group(1))
        except ValueError:
            continue
        if parsed <= trade_date:
            dated.append((parsed, path))
    if not dated:
        return None, pd.DataFrame()
    _, path = max(dated, key=lambda item: item[0])
    return path, pd.read_csv(path)


def _read_optional_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path)


def _previous_pool_candidates(payload: dict[str, object], *, limit: int | None) -> list[dict[str, object]]:
    raw_candidates = payload.get("candidates") if isinstance(payload, dict) else []
    if not isinstance(raw_candidates, list):
        return []

    extracted_symbols = _extract_pool_symbols(payload)
    order = {symbol: index for index, symbol in enumerate(extracted_symbols)}
    candidates: list[dict[str, object]] = []
    seen: set[str] = set()
    for raw in raw_candidates:
        if not isinstance(raw, dict):
            continue
        symbol = normalize_symbol(raw.get("symbol", ""))
        if not symbol or symbol in seen:
            continue
        seen.add(symbol)
        candidates.append(
            {
                "symbol": symbol,
                "name": str(raw.get("name", "") or ""),
                "prev_rank": order.get(symbol, len(order)) + 1,
                "prev_source": raw.get("source"),
                "prev_source_tags": _jsonish(raw.get("source_tags")),
                "prev_pattern_match": raw.get("pattern_match"),
                "prev_pattern_id": raw.get("pattern_id"),
                "prev_pattern_ids": _jsonish(raw.get("pattern_ids")),
                "prev_patterns": _jsonish(raw.get("patterns")),
                "prev_reason": raw.get("reason"),
                "prev_phase1_score_100": raw.get("phase1_score_100"),
                "prev_phase2_score_100": raw.get("phase2_score_100"),
                "prev_phase4_score_100": raw.get("phase4_score_100"),
                "prev_phase5_score_100": raw.get("phase5_score_100"),
                "prev_watchlist_streak": raw.get("连续上榜天数"),
            }
        )
    candidates = sorted(candidates, key=lambda item: int(item["prev_rank"]))
    if limit is not None:
        candidates = candidates[: max(int(limit), 0)]
    return candidates


def _extract_pool_symbols(payload: dict[str, object]) -> list[str]:
    candidates = payload.get("candidates")
    if not isinstance(candidates, list):
        return []
    symbols: list[str] = []
    seen: set[str] = set()
    for item in candidates:
        if not isinstance(item, dict):
            continue
        symbol = normalize_symbol(item.get("symbol", ""))
        if not symbol or symbol in seen:
            continue
        symbols.append(symbol)
        seen.add(symbol)
    return symbols


def _candidate_universe(candidates: list[dict[str, object]]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "symbol": candidate["symbol"],
                "name": candidate.get("name", ""),
                "latest_price": pd.NA,
                "volume": pd.NA,
                "amount": pd.NA,
                "turnover_rate": pd.NA,
            }
            for candidate in candidates
        ]
    )


def _load_tracked_symbols(project_root: Path) -> list[str]:
    path = project_root / DEFAULT_TRACK_STOCK_FILENAME
    if not path.exists():
        return []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except OSError:
        return []
    try:
        if TRACK_INPUT_SHEET not in workbook.sheetnames:
            return []
        sheet = workbook[TRACK_INPUT_SHEET]
        header_values = [str(cell.value or "").strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        symbol_column = 1
        has_symbol_header = False
        for index, value in enumerate(header_values, start=1):
            if value in {"symbol", "code", "股票代码", "代码", "证券代码"}:
                symbol_column = index
                has_symbol_header = True
                break
        start_row = 2 if has_symbol_header else 1
        symbols: list[str] = []
        seen: set[str] = set()
        for row in sheet.iter_rows(min_row=start_row, min_col=symbol_column, max_col=symbol_column, values_only=True):
            symbol = normalize_symbol(row[0] if row else "")
            if not symbol or symbol in seen:
                continue
            symbols.append(symbol)
            seen.add(symbol)
        return symbols
    finally:
        workbook.close()


def _merge_tracked_candidates(
    *,
    storage: Storage,
    candidates: list[dict[str, object]],
    tracked_symbols: list[str],
) -> list[dict[str, object]]:
    tracked = {normalize_symbol(symbol) for symbol in tracked_symbols if normalize_symbol(symbol)}
    if not tracked:
        return candidates
    result = [dict(candidate) for candidate in candidates]
    by_symbol = {str(candidate.get("symbol", "")): candidate for candidate in result}
    for candidate in result:
        candidate["track_stock"] = str(candidate.get("symbol", "")) in tracked

    missing = [symbol for symbol in tracked_symbols if normalize_symbol(symbol) not in by_symbol]
    if not missing:
        return result

    universe_names = _universe_name_lookup(storage)
    for symbol in missing:
        normalized = normalize_symbol(symbol)
        if not normalized:
            continue
        result.append(
            {
                "symbol": normalized,
                "name": universe_names.get(normalized, ""),
                "prev_rank": pd.NA,
                "universe_rank": pd.NA,
                "prev_source": "",
                "prev_source_tags": "",
                "prev_pattern_match": False,
                "prev_pattern_id": "",
                "prev_pattern_ids": "",
                "prev_patterns": "",
                "prev_reason": "",
                "prev_phase1_score_100": pd.NA,
                "prev_phase2_score_100": pd.NA,
                "prev_phase4_score_100": pd.NA,
                "prev_phase5_score_100": pd.NA,
                "prev_watchlist_streak": pd.NA,
                "track_stock": True,
            }
        )
    return result


def _universe_name_lookup(storage: Storage) -> dict[str, str]:
    try:
        universe = storage.load_universe().copy()
    except FileNotFoundError:
        return {}
    if "symbol" not in universe.columns:
        return {}
    universe["symbol"] = universe["symbol"].map(normalize_symbol)
    return {
        str(row.get("symbol", "")): str(row.get("name", "") or "")
        for row in universe.to_dict("records")
        if str(row.get("symbol", ""))
    }


def _run_candidate_analysis(
    *,
    storage: Storage,
    project_root: Path,
    trade_date: date,
    candidates: list[dict[str, object]],
    output_dir: Path,
    file_tag: str,
) -> _AnalysisResult:
    symbols = [candidate["symbol"] for candidate in candidates]
    universe = _candidate_universe(candidates)
    overlay_storage = _IntradayOverlayStorage(storage=storage, universe=universe, trade_date=trade_date)
    phase1_path = output_dir / f"intraday_{file_tag}_tail_risk_predictions_{trade_date.isoformat()}.csv"
    phase2_path = output_dir / f"intraday_{file_tag}_barrier_risk_predictions_{trade_date.isoformat()}.csv"
    phase4_path = output_dir / f"intraday_{file_tag}_alpha158_qlib_return_predictions_{trade_date.isoformat()}.csv"
    phase8_path = output_dir / f"intraday_{file_tag}_limit_up_3d_opportunity_predictions_{trade_date.isoformat()}.csv"

    phase1 = predict_tail_risk(
        storage=overlay_storage,
        project_root=project_root,
        trade_date=trade_date,
        output=phase1_path,
        latest_only=True,
        feature_lookback_bars=61,
        include_features=False,
        prediction_scope="intraday_screening",
    ).predictions
    phase2 = predict_barrier_risk(
        storage=overlay_storage,
        project_root=project_root,
        trade_date=trade_date,
        output=phase2_path,
        latest_only=True,
        feature_lookback_bars=61,
        include_features=False,
        prediction_scope="intraday_screening",
    ).predictions
    phase4 = predict_alpha158_qlib_return(
        storage=overlay_storage,
        project_root=project_root,
        trade_date=trade_date,
        output=phase4_path,
        latest_only=True,
        feature_lookback_bars=61,
        include_features=False,
        prediction_scope="intraday_screening",
    ).predictions
    phase8 = _predict_intraday_phase8_if_available(
        storage=overlay_storage,
        project_root=project_root,
        trade_date=trade_date,
        output=phase8_path,
    )

    macd = _build_intraday_macd_summary(overlay_storage, trade_date=trade_date, symbols=symbols)
    atr = _build_intraday_atr_summary(overlay_storage, trade_date=trade_date, symbols=symbols)
    intraday = _load_intraday_snapshot_frame(storage, symbols=symbols)
    frame = _build_output_frame(
        project_root=project_root,
        trade_date=trade_date,
        candidates=candidates,
        intraday=intraday,
        phase1=phase1,
        phase2=phase2,
        phase4=phase4,
        phase8=phase8,
        macd=macd,
        atr=atr,
    )
    return _AnalysisResult(
        frame=frame,
        candidates=candidates,
        intraday=intraday,
        phase1=phase1,
        phase2=phase2,
        phase4=phase4,
        phase8=phase8,
        macd=macd,
        atr=atr,
        phase1_path=phase1_path,
        phase2_path=phase2_path,
        phase4_path=phase4_path,
        phase8_path=phase8_path if not phase8.empty else None,
    )


def _concat_frames(*frames: pd.DataFrame) -> pd.DataFrame:
    available = [frame for frame in frames if frame is not None and not frame.empty]
    if not available:
        return pd.DataFrame()
    return pd.concat(available, ignore_index=True)


def _predict_intraday_phase8_if_available(
    *,
    storage: _IntradayOverlayStorage,
    project_root: Path,
    trade_date: date,
    output: Path,
) -> pd.DataFrame:
    if not limit_up_3d_model_path(project_root).exists():
        return pd.DataFrame(columns=["symbol"])
    return predict_limit_up_3d_opportunity(
        storage=storage,
        project_root=project_root,
        trade_date=trade_date,
        output=output,
        latest_only=True,
        feature_lookback_bars=61,
        include_features=False,
        prediction_scope="intraday_screening",
    ).predictions


def _save_intraday_track_stock(*, output_dir: Path, trade_date: date, frame: pd.DataFrame) -> Path | None:
    track = _select_intraday_track_stock(frame)
    if track.empty:
        return None
    target = output_dir / f"intraday_track_stock_{trade_date.isoformat()}.csv"
    track.to_csv(target, index=False, encoding="utf-8-sig")
    return target


def _select_intraday_track_stock(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty or "track_stock" not in frame.columns:
        return frame.head(0).copy()
    result = frame[frame["track_stock"].fillna(False).astype(bool)].copy()
    if result.empty:
        return result
    result = _add_intraday_pool_score(result)
    result = _sort_intraday_pool(result)
    return result.reset_index(drop=True)


def _add_intraday_pool_score(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    required = {"phase1_score_100", "phase2_score_100", "phase4_score_100"}
    if not required.issubset(result.columns):
        result["intraday_pool_score"] = pd.NA
        return result
    result = add_centered_risk_scores(result)
    result["intraday_pool_score"] = result["centered_risk_score"]
    return result


def _add_intraday_selection_source(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    base_source = result.get("prev_source", pd.Series("", index=result.index)).fillna("").astype(str)
    source = base_source.where(base_source.str.strip().ne(""), "intraday_pool")
    if "track_stock" in result.columns:
        tracked = result["track_stock"].fillna(False).astype(bool)
        source = source.mask(tracked & base_source.str.strip().ne(""), source + "+track_stock")
        source = source.mask(tracked & base_source.str.strip().eq(""), "track_stock")
    result["intraday_selection_source"] = source
    return result


def _sort_intraday_pool(
    frame: pd.DataFrame,
    *,
    leading_columns: list[str] | None = None,
    leading_ascending: list[bool] | None = None,
) -> pd.DataFrame:
    result = _add_intraday_pool_score(frame)
    leading = leading_columns or []
    leading_directions = leading_ascending if leading_ascending is not None else [True] * len(leading)
    sort_columns = [
        *leading,
        "intraday_pool_score",
        "phase4_score_100",
        "phase1_center_score",
        "phase2_center_score",
        "symbol",
    ]
    ascending = [
        *leading_directions,
        False,
        False,
        False,
        False,
        True,
    ]
    available_columns: list[str] = []
    available_ascending: list[bool] = []
    for column, direction in zip(sort_columns, ascending):
        if column in result.columns:
            available_columns.append(column)
            available_ascending.append(direction)
    if not available_columns:
        return result
    return result.sort_values(available_columns, ascending=available_ascending, na_position="last")


def _cleanup_intraday_screening_reports(
    output_dir: Path,
    *,
    keep_dates: int | None,
    preserve_paths: tuple[Path | None, ...],
) -> int:
    preserved = {_resolved(path) for path in preserve_paths if path is not None}
    deleted = 0
    for path in output_dir.glob("intraday_*_predictions_*.csv"):
        if _delete_report_file(path, preserved):
            deleted += 1

    if keep_dates is None:
        return deleted

    keep_count = max(int(keep_dates), 1)
    dated_files: list[tuple[date, Path]] = []
    for path in output_dir.iterdir():
        if not path.is_file():
            continue
        report_date = _dated_intraday_report_date(path.name)
        if report_date is not None:
            dated_files.append((report_date, path))
    keep = set(sorted({report_date for report_date, _ in dated_files}, reverse=True)[:keep_count])
    for report_date, path in dated_files:
        if report_date not in keep and _delete_report_file(path, preserved):
            deleted += 1
    return deleted


def _dated_intraday_report_date(filename: str) -> date | None:
    for pattern in _DATED_REPORT_PATTERNS:
        match = pattern.match(filename)
        if match is None:
            continue
        try:
            return date.fromisoformat(match.group(1))
        except ValueError:
            return None
    return None


def _delete_report_file(path: Path, preserved: set[Path]) -> bool:
    if not path.is_file() or _resolved(path) in preserved:
        return False
    try:
        path.unlink()
    except OSError:
        return False
    return True


def _resolved(path: Path) -> Path:
    return path.resolve()


def _normalize_intraday_bar_for_daily(intraday: pd.DataFrame, *, symbol: str, trade_date: date) -> pd.DataFrame:
    if intraday.empty:
        return pd.DataFrame()
    frame = intraday.copy()
    frame["trade_date"] = pd.to_datetime(frame.get("trade_date"), errors="coerce")
    frame = frame[frame["trade_date"].dt.date.eq(trade_date)].copy()
    if frame.empty:
        return pd.DataFrame()
    if "quote_datetime" in frame.columns:
        frame["quote_datetime"] = pd.to_datetime(frame["quote_datetime"], errors="coerce")
        frame = frame.sort_values("quote_datetime", na_position="first")
    latest = frame.tail(1).copy()
    latest["symbol"] = symbol
    for column in ("open", "high", "low", "close", "pre_close", "volume", "amount"):
        latest[column] = pd.to_numeric(latest.get(column), errors="coerce")
    latest["change"] = latest["close"].sub(latest["pre_close"])
    latest["pct_change"] = latest["close"].div(latest["pre_close"]).sub(1.0).mul(100.0)
    latest["amplitude"] = latest["high"].sub(latest["low"]).div(latest["pre_close"].replace(0, pd.NA)).mul(100.0)
    latest["turnover"] = float("nan")
    keep = [
        "trade_date",
        "symbol",
        "open",
        "close",
        "high",
        "low",
        "volume",
        "amount",
        "pct_change",
        "change",
        "amplitude",
        "turnover",
    ]
    return latest.loc[:, keep].reset_index(drop=True)


def _build_intraday_macd_summary(storage: _IntradayOverlayStorage, *, trade_date: date, symbols: list[str]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    universe = storage.load_universe().set_index("symbol")
    for symbol in symbols:
        try:
            bars = storage.load_daily_bars(symbol)
        except (FileNotFoundError, DailyBarsReadError):
            continue
        cutoff = bars[pd.to_datetime(bars["trade_date"], errors="coerce").dt.date <= trade_date].reset_index(drop=True)
        if cutoff.empty:
            continue
        macd_frame = _prepare_daily_macd_frame(cutoff)
        latest = macd_frame.iloc[-1]
        divergence = summarize_recent_macd_divergence(macd_frame)
        bullish_volume_divergence, bearish_volume_divergence = _detect_daily_volume_price_divergence(macd_frame)
        rows.append(
            {
                "trade_date": trade_date.isoformat(),
                "symbol": symbol,
                "name": str(universe.loc[symbol].get("name", "")) if symbol in universe.index else "",
                "macd": _safe_float_or_none(latest.get("macd")),
                "macd_signal_line": _safe_float_or_none(latest.get("macd_signal_line")),
                "macd_hist": _safe_float_or_none(latest.get("macd_hist")),
                "macd_cross_state": _describe_macd_cross_state(macd_frame),
                "macd_divergence_state": _describe_macd_divergence_state(divergence),
                "volume_price_divergence_state": _describe_volume_price_divergence_state(
                    bullish_volume_divergence,
                    bearish_volume_divergence,
                ),
                "macd_top_divergence_15d": bool(divergence.get("macd_top_divergence_15d", False)),
                "macd_bottom_divergence_15d": bool(divergence.get("macd_bottom_divergence_15d", False)),
                "macd_top_divergence_signal_date": divergence.get("macd_top_divergence_signal_date"),
                "macd_bottom_divergence_signal_date": divergence.get("macd_bottom_divergence_signal_date"),
                "bullish_volume_price_divergence_flag": bool(bullish_volume_divergence),
                "bearish_volume_price_divergence_flag": bool(bearish_volume_divergence),
            }
        )
    return pd.DataFrame(rows)


def _build_intraday_atr_summary(storage: _IntradayOverlayStorage, *, trade_date: date, symbols: list[str]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    universe = storage.load_universe().set_index("symbol")
    for symbol in symbols:
        try:
            bars = storage.load_daily_bars(symbol)
        except (FileNotFoundError, DailyBarsReadError):
            continue
        cutoff = bars[pd.to_datetime(bars["trade_date"], errors="coerce").dt.date <= trade_date].reset_index(drop=True)
        if cutoff.empty:
            continue
        snapshot = build_atr_snapshot_row(
            cutoff,
            symbol=symbol,
            name=str(universe.loc[symbol].get("name", "")) if symbol in universe.index else "",
            trade_date=trade_date,
        )
        if snapshot is not None:
            rows.append(snapshot)
    return pd.DataFrame(rows)


def _load_intraday_snapshot_frame(storage: Storage, *, symbols: list[str]) -> pd.DataFrame:
    rows: list[pd.DataFrame] = []
    for symbol in symbols:
        try:
            frame = storage.load_intraday_bars(symbol)
        except FileNotFoundError:
            continue
        if frame.empty:
            continue
        copied = frame.copy()
        copied["symbol"] = copied["symbol"].map(normalize_symbol)
        if "quote_datetime" in copied.columns:
            copied["quote_datetime"] = pd.to_datetime(copied["quote_datetime"], errors="coerce")
            copied = copied.sort_values("quote_datetime", na_position="first")
        rows.append(copied.tail(1))
    if not rows:
        return pd.DataFrame(columns=["symbol"])
    result = pd.concat(rows, ignore_index=True)
    for column in ("trade_date", "quote_datetime"):
        if column in result.columns:
            result[column] = pd.to_datetime(result[column], errors="coerce")
    if "pct_change" not in result.columns:
        result["pct_change"] = pd.NA
    pct_change = pd.to_numeric(result["pct_change"], errors="coerce")
    if {"close", "pre_close"}.issubset(result.columns):
        close = pd.to_numeric(result["close"], errors="coerce")
        pre_close = pd.to_numeric(result["pre_close"], errors="coerce").replace(0, pd.NA)
        derived_pct_change = close.div(pre_close).sub(1.0).mul(100.0)
        result["pct_change"] = pct_change.where(pct_change.notna(), derived_pct_change)
    return result


def _build_output_frame(
    *,
    project_root: Path,
    trade_date: date,
    candidates: list[dict[str, object]],
    intraday: pd.DataFrame,
    phase1: pd.DataFrame,
    phase2: pd.DataFrame,
    phase4: pd.DataFrame,
    phase8: pd.DataFrame,
    macd: pd.DataFrame,
    atr: pd.DataFrame,
) -> pd.DataFrame:
    base = pd.DataFrame(candidates)
    phase1_prepared = _prepare_phase_risk_predictions(
        phase1,
        score_column="risk_score",
        output_score_column="phase1_risk_score",
        prefix="phase1",
        filter_rate=0.2,
        model_prefix="phase1",
    )
    phase2_prepared = _prepare_phase_risk_predictions(
        phase2,
        score_column="barrier_risk_score",
        output_score_column="phase2_barrier_risk_score",
        prefix="phase2",
        filter_rate=0.2,
        model_prefix="phase2",
        extra_columns=("is_cusum_event", "mlfin_daily_vol", "mlfin_cusum_threshold"),
    )
    phase4_prepared = _prepare_phase4_predictions(phase4)
    phase4_prepared = merge_phase4_rolling_frame(phase4_prepared, project_root=project_root, trade_date=trade_date)
    phase8_prepared = _prepare_phase8_predictions(phase8)
    macd_prepared = _prepare_macd_frame(macd)
    atr_prepared = _prepare_atr_frame(atr).rename(columns={"trade_date": "atr_trade_date", "close": "atr_close"})
    intraday_prepared = _prepare_intraday_for_output(intraday)

    result = base.copy()
    for frame in (intraday_prepared, phase1_prepared, phase2_prepared, phase4_prepared, phase8_prepared, macd_prepared, atr_prepared):
        if frame.empty or "symbol" not in frame.columns:
            continue
        frame = frame.drop(columns=["name"], errors="ignore")
        result = result.merge(frame, on="symbol", how="left")
    result = _add_phase_display_ranks(result)
    if "phase5_score_100" not in result.columns:
        result["phase5_score_100"] = result.get("prev_phase5_score_100", pd.NA)
    else:
        result["phase5_score_100"] = result["phase5_score_100"].where(
            result["phase5_score_100"].notna(),
            result.get("prev_phase5_score_100", pd.NA),
        )
    result = _add_intraday_pool_score(result)
    result = _add_intraday_selection_source(result)
    result = add_recommended_position_percent(result)
    result = _drop_internal_score_columns(result)
    return _order_output_columns(result)


def _prepare_intraday_for_output(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty or "symbol" not in frame.columns:
        return pd.DataFrame(columns=["symbol"])
    result = frame.copy()
    result["symbol"] = result["symbol"].map(normalize_symbol)
    rename_map = {
        "trade_date": "intraday_trade_date",
        "pct_change": "intraday_pct_change",
        "quote_datetime": "intraday_quote_datetime",
        "quote_time": "intraday_quote_time",
        "source": "intraday_source",
        "fetched_at": "intraday_fetched_at",
        "provisional": "intraday_provisional",
    }
    result = result.rename(columns={source: target for source, target in rename_map.items() if source in result.columns})
    for column in ("intraday_trade_date", "intraday_quote_datetime"):
        if column in result.columns:
            result[column] = pd.to_datetime(result[column], errors="coerce").astype("string")
    keep = ["symbol", *rename_map.values()]
    return result.loc[:, [column for column in keep if column in result.columns]].drop_duplicates("symbol", keep="last")


def _add_phase_display_ranks(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    for phase in ("phase1", "phase2", "phase4"):
        score_column = f"{phase}_score_100"
        rank_column = f"{phase}_rank"
        if score_column not in result.columns:
            result[rank_column] = pd.NA
            continue
        scores = pd.to_numeric(result[score_column], errors="coerce")
        result[rank_column] = scores.rank(method="min", ascending=False).astype("Int64")
    return result


def _drop_internal_score_columns(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.drop(
        columns=[
            "phase1_risk_score",
            "phase1_risk_rank",
            "phase1_risk_percentile",
            "phase2_barrier_risk_score",
            "phase2_risk_rank",
            "phase2_risk_percentile",
            "phase4_return_score",
            "phase4_score_percentile",
        ],
        errors="ignore",
    )


def _order_output_columns(frame: pd.DataFrame) -> pd.DataFrame:
    technical_columns = [
        "macd_cross_state",
        "macd_divergence_state",
        "volume_price_divergence_state",
        "macd_top_divergence_15d",
        "macd_bottom_divergence_15d",
        "macd_top_divergence_signal_date",
        "macd_bottom_divergence_signal_date",
        "bullish_volume_price_divergence_flag",
        "bearish_volume_price_divergence_flag",
        "macd",
        "macd_signal_line",
        "macd_hist",
        "atr_trade_date",
        "atr_close",
        "atr_14",
        "atr_stop_loss_1x",
        "atr_stop_loss_2x",
        "atr_take_profit_2x",
        "atr_take_profit_3x",
        "atr_volatility_regime",
    ]
    phase_detail_columns = [
        "intraday_pool_score",
        "phase4_5d_std",
        "centered_risk_score",
        "phase1_center_score",
        "phase2_center_score",
        "phase4_rank",
        *PHASE4_ROLLING_RANK_COLUMNS,
        "phase8_rank",
        "phase8_raw_score",
        "phase8_feature_trade_date",
        "phase8_model_name",
        "phase8_model_version",
        "today_limit_up_excluded",
        "today_high_return_vs_prev_close",
        "today_close_return_vs_prev_close",
        "phase1_rank",
        "phase2_rank",
        "phase1_excluded_by_top20_risk",
        "phase2_excluded_by_top20_risk",
        "phase2_is_cusum_event",
        "phase1_feature_trade_date",
        "phase2_feature_trade_date",
        "phase4_feature_trade_date",
        "phase1_model_name",
        "phase1_model_version",
        "phase2_mlfin_daily_vol",
        "phase2_mlfin_cusum_threshold",
        "phase2_model_name",
        "phase2_model_version",
        "phase4_name",
        "phase4_model_name",
        "phase4_model_version",
        "track_stock",
        "prev_rank",
        "universe_rank",
        "intraday_quote_datetime",
        "intraday_quote_time",
        "intraday_fetched_at",
        "intraday_provisional",
    ]
    pattern_detail_columns = [
        "prev_source_tags",
        "prev_reason",
        "prev_watchlist_streak",
        "prev_patterns",
        "prev_phase1_score_100",
        "prev_phase2_score_100",
        "prev_phase4_score_100",
        "prev_phase5_score_100",
    ]
    preferred = [
        "intraday_trade_date",
        "symbol",
        "name",
        "intraday_selection_source",
        "intraday_pct_change",
        "phase1_score_100",
        "phase2_score_100",
        "phase4_score_100",
        "phase8_score_100",
        "phase4_5d_mean",
        "prev_pattern_match",
        "prev_pattern_ids",
        "prev_pattern_id",
        "atr_pct_14",
        RECOMMENDED_POSITION_PERCENT_FIELD,
        *technical_columns,
        "prev_source",
        "phase5_score_100",
        "intraday_source",
        *phase_detail_columns,
        *pattern_detail_columns,
    ]
    seen_columns: set[str] = set()
    columns = []
    for column in preferred:
        if column in frame.columns and column not in seen_columns:
            columns.append(column)
            seen_columns.add(column)
    columns.extend([column for column in frame.columns if column not in columns])
    ordered = frame.loc[:, columns].copy()
    sort_columns = [column for column in ("prev_rank", "phase4_rank", "phase1_rank", "phase2_rank", "symbol") if column in ordered.columns]
    if sort_columns:
        ordered = ordered.sort_values(sort_columns, ascending=[True] * len(sort_columns), na_position="last")
    return ordered.reset_index(drop=True)


def _prepare_daily_macd_frame(dataframe: pd.DataFrame) -> pd.DataFrame:
    frame = dataframe.copy().sort_values("trade_date").reset_index(drop=True)
    if not {"macd_dif", "macd_dea", "macd_hist"}.issubset(frame.columns):
        frame = add_indicators(frame).sort_values("trade_date").reset_index(drop=True)
    if "macd" not in frame.columns and "macd_dif" in frame.columns:
        frame["macd"] = frame["macd_dif"]
    if "macd_signal_line" not in frame.columns and "macd_dea" in frame.columns:
        frame["macd_signal_line"] = frame["macd_dea"]
    return frame


def _describe_macd_cross_state(dataframe: pd.DataFrame) -> str:
    normalized = _prepare_daily_macd_frame(dataframe)
    if normalized.empty or "macd" not in normalized.columns or "macd_signal_line" not in normalized.columns:
        return "unknown"
    recent = normalized.tail(3).reset_index(drop=True)
    recent_cross_up = False
    recent_cross_down = False
    for offset in range(1, len(recent)):
        prev_row = recent.iloc[offset - 1]
        current_row = recent.iloc[offset]
        if pd.isna(prev_row.get("macd")) or pd.isna(prev_row.get("macd_signal_line")):
            continue
        if pd.isna(current_row.get("macd")) or pd.isna(current_row.get("macd_signal_line")):
            continue
        if float(prev_row["macd"]) <= float(prev_row["macd_signal_line"]) and float(current_row["macd"]) > float(current_row["macd_signal_line"]):
            recent_cross_up = True
        if float(prev_row["macd"]) >= float(prev_row["macd_signal_line"]) and float(current_row["macd"]) < float(current_row["macd_signal_line"]):
            recent_cross_down = True
    latest = recent.iloc[-1]
    macd = _safe_float_or_none(latest.get("macd"))
    signal_line = _safe_float_or_none(latest.get("macd_signal_line"))
    if recent_cross_up:
        return "golden_cross"
    if recent_cross_down:
        return "dead_cross"
    if macd is None or signal_line is None:
        return "unknown"
    return "above_signal" if macd >= signal_line else "below_signal"


def _detect_daily_volume_price_divergence(dataframe: pd.DataFrame) -> tuple[bool, bool]:
    if dataframe.empty or len(dataframe) < 6:
        return False, False
    frame = dataframe.copy().sort_values("trade_date").reset_index(drop=True)
    recent = frame.tail(5).reset_index(drop=True)
    previous = frame.iloc[-6]
    latest = recent.iloc[-1]
    recent_avg_volume = pd.to_numeric(recent["volume"], errors="coerce").mean()
    previous_volume = _safe_float_or_none(previous.get("volume"))
    latest_close = _safe_float_or_none(latest.get("close"))
    previous_close = _safe_float_or_none(previous.get("close"))
    if previous_volume is None or latest_close is None or previous_close is None or pd.isna(recent_avg_volume):
        return False, False
    bullish = latest_close > previous_close and float(recent_avg_volume) < previous_volume * 0.9
    bearish = latest_close < previous_close and float(recent_avg_volume) > previous_volume * 1.1
    return bullish, bearish


def _describe_macd_divergence_state(macd_summary: dict[str, object]) -> str:
    if bool(macd_summary.get("macd_bottom_divergence_15d", False)):
        return "bottom_divergence"
    if bool(macd_summary.get("macd_top_divergence_15d", False)):
        return "top_divergence"
    return "none"


def _describe_volume_price_divergence_state(bullish: bool, bearish: bool) -> str:
    if bullish and not bearish:
        return "bullish"
    if bearish and not bullish:
        return "bearish"
    return "none"


def _safe_float_or_none(value: object) -> float | None:
    if value is None or pd.isna(value):
        return None
    return float(value)


def _jsonish(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)
