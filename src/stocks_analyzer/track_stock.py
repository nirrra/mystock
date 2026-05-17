from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import math

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .daily_returns import read_full_market_daily_returns, write_full_market_daily_returns
from .phase_display import add_phase5_score_100, phase7_score_100, score_series_100
from .phase4_rolling import merge_phase4_rolling_frame
from .position_sizing import RECOMMENDED_POSITION_PERCENT_FIELD, recommended_position_percent_from_mapping
from .sector_membership import build_sector_display_frame


DEFAULT_TRACK_STOCK_FILENAME = "track_stock.xlsx"
TRACK_INPUT_SHEET = "Sheet1"
TRACK_DAILY_OUTPUT_SHEET = "Sheet2"
TRACK_INTRADAY_OUTPUT_SHEET = "Sheet3"
TRACK_OUTPUT_SHEET = TRACK_DAILY_OUTPUT_SHEET
TEXT_NUMBER_FORMAT = "@"


@dataclass(frozen=True)
class TrackStockUpdateResult:
    workbook_path: Path
    trade_date: date
    tracked_count: int
    output_rows: int
    output_sheet: str


TRACK_STOCK_COLUMNS = [
    "trade_date",
    "symbol",
    "name",
    "daily_return_pct",
    "daily_close",
    "industry_names",
    "concept_names",
    "phase1_score_100",
    "phase2_score_100",
    "phase2_is_cusum_event",
    "phase4_score_100",
    "phase4_5d_mean",
    "phase4_5d_std",
    "phase5_score_100",
    "phase8_score_100",
    "phase7_score_100",
    "phase7_trade_permission",
    "pattern_match",
    "pattern_ids",
    "pattern1",
    "pattern2",
    "pattern3",
    "pattern4",
    "pattern5",
    "pattern6",
    "macd",
    "macd_signal_line",
    "macd_hist",
    "macd_cross_state",
    "macd_divergence_state",
    "volume_price_divergence_state",
    "macd_top_divergence_15d",
    "macd_bottom_divergence_15d",
    "bullish_volume_price_divergence_flag",
    "bearish_volume_price_divergence_flag",
    "atr_14",
    "atr_pct_14",
    "atr_stop_loss_1x",
    "atr_stop_loss_2x",
    "atr_take_profit_2x",
    "atr_take_profit_3x",
    "recommended_position_pct",
    "atr_volatility_regime",
]

TRACK_STOCK_HEADERS_ZH = {
    "trade_date": "交易日期",
    "symbol": "股票代码",
    "name": "股票名称",
    "daily_return_pct": "当日涨幅%",
    "daily_close": "收盘价",
    "industry_names": "所属行业",
    "concept_names": "所属概念",
    "phase1_score_100": "Phase1买入分",
    "phase2_score_100": "Phase2买入分",
    "phase2_is_cusum_event": "Phase2是否CUSUM事件",
    "phase4_score_100": "Phase4买入分",
    "phase4_5d_mean": "Phase4五日均分",
    "phase4_5d_std": "Phase4五日STD",
    "phase5_score_100": "Phase5买入分",
    "phase8_score_100": "Phase8短线分",
    "phase7_score_100": "Phase7交易日分",
    "phase7_trade_permission": "Phase7交易许可",
    "pattern_match": "是否命中模式",
    "pattern_ids": "命中模式",
    "pattern1": "模式1",
    "pattern2": "模式2",
    "pattern3": "模式3",
    "pattern4": "模式4",
    "pattern5": "模式5",
    "pattern6": "模式6",
    "macd": "MACD",
    "macd_signal_line": "MACD信号线",
    "macd_hist": "MACD柱",
    "macd_cross_state": "MACD金叉死叉",
    "macd_divergence_state": "MACD背离",
    "volume_price_divergence_state": "量价背离",
    "macd_top_divergence_15d": "15日MACD顶背离",
    "macd_bottom_divergence_15d": "15日MACD底背离",
    "bullish_volume_price_divergence_flag": "量价底背离标记",
    "bearish_volume_price_divergence_flag": "量价顶背离标记",
    "atr_14": "ATR14",
    "atr_pct_14": "ATR%",
    "atr_stop_loss_1x": "1ATR止损",
    "atr_stop_loss_2x": "2ATR止损",
    "atr_take_profit_2x": "2ATR止盈",
    "atr_take_profit_3x": "3ATR止盈",
    "recommended_position_pct": RECOMMENDED_POSITION_PERCENT_FIELD,
    "atr_volatility_regime": "ATR波动分层",
}

INTRADAY_TRACK_STOCK_COLUMNS = [
    "intraday_trade_date",
    "symbol",
    "name",
    "intraday_selection_source",
    "intraday_pct_change",
    "phase1_score_100",
    "phase2_score_100",
    "phase4_score_100",
    "phase4_5d_mean",
    "phase4_5d_std",
    "intraday_pool_score",
    "prev_pattern_match",
    "prev_pattern_ids",
    "atr_pct_14",
    "recommended_position_pct",
    "atr_close",
    "atr_14",
    "macd_cross_state",
    "macd_divergence_state",
    "volume_price_divergence_state",
    "industry_names",
    "concept_names",
    "intraday_quote_datetime",
    "intraday_source",
]

INTRADAY_TRACK_STOCK_HEADERS_ZH = {
    "intraday_trade_date": "交易日期",
    "symbol": "股票代码",
    "name": "股票名称",
    "intraday_selection_source": "来源",
    "intraday_pct_change": "盘中涨幅%",
    "phase1_score_100": "P1风险质量分",
    "phase2_score_100": "P2交易风险分",
    "phase4_score_100": "P4上涨质量分",
    "phase4_5d_mean": "P4五日均分",
    "phase4_5d_std": "P4五日std",
    "intraday_pool_score": "P1/P2/P4综合分",
    "prev_pattern_match": "Pattern命中",
    "prev_pattern_ids": "Pattern编号",
    "atr_pct_14": "ATR%",
    "recommended_position_pct": RECOMMENDED_POSITION_PERCENT_FIELD,
    "atr_close": "最新价格",
    "atr_14": "ATR14",
    "macd_cross_state": "macd交叉",
    "macd_divergence_state": "macd背离",
    "volume_price_divergence_state": "量价背离",
    "industry_names": "行业",
    "concept_names": "概念",
    "intraday_quote_datetime": "行情时间",
    "intraday_source": "行情源",
}

INTRADAY_CHINESE_TO_INTERNAL = {
    "日期": "intraday_trade_date",
    "交易日期": "intraday_trade_date",
    "编号": "symbol",
    "股票代码": "symbol",
    "股票名称": "name",
    "来源": "intraday_selection_source",
    "盘中涨幅%": "intraday_pct_change",
    "P1风险质量分": "phase1_score_100",
    "P2交易风险分": "phase2_score_100",
    "P4上涨质量分": "phase4_score_100",
    "P4五日均分": "phase4_5d_mean",
    "P4五日std": "phase4_5d_std",
    "P1/P2/P4综合分": "intraday_pool_score",
    "Pattern命中": "prev_pattern_match",
    "Pattern编号": "prev_pattern_ids",
    "ATR%": "atr_pct_14",
    RECOMMENDED_POSITION_PERCENT_FIELD: "recommended_position_pct",
    "最新价格": "atr_close",
    "ATR14": "atr_14",
    "macd交叉": "macd_cross_state",
    "macd背离": "macd_divergence_state",
    "量价背离": "volume_price_divergence_state",
    "行业": "industry_names",
    "概念": "concept_names",
    "行情时间": "intraday_quote_datetime",
    "行情源": "intraday_source",
}


def update_track_stock_workbook(
    *,
    project_root: Path,
    trade_date: date,
    workbook_path: Path | None = None,
    mode: str = "daily",
    intraday_frame: pd.DataFrame | None = None,
) -> TrackStockUpdateResult:
    target = _resolve_workbook_path(project_root, workbook_path)
    workbook = _load_or_create_workbook(target)
    symbols = _read_tracked_symbols(workbook)

    normalized_mode = str(mode or "daily").strip().lower()
    if normalized_mode in {"daily", "post", "postmarket", "盘后"}:
        rows = _build_daily_rows(project_root=project_root, trade_date=trade_date, symbols=symbols)
        output_sheet = TRACK_DAILY_OUTPUT_SHEET
        columns = TRACK_STOCK_COLUMNS
        headers = TRACK_STOCK_HEADERS_ZH
    elif normalized_mode in {"intraday", "盘中"}:
        rows = _build_intraday_rows(
            project_root=project_root,
            trade_date=trade_date,
            symbols=symbols,
            intraday_frame=intraday_frame,
        )
        output_sheet = TRACK_INTRADAY_OUTPUT_SHEET
        columns = INTRADAY_TRACK_STOCK_COLUMNS
        headers = INTRADAY_TRACK_STOCK_HEADERS_ZH
    else:
        raise ValueError(f"Unsupported track-stock mode: {mode}")

    _write_output_sheet(workbook, rows, sheet_name=output_sheet, columns=columns, headers=headers)
    _ensure_sheet_order(workbook)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return TrackStockUpdateResult(
        workbook_path=target,
        trade_date=trade_date,
        tracked_count=len(symbols),
        output_rows=len(rows),
        output_sheet=output_sheet,
    )


def _build_daily_rows(*, project_root: Path, trade_date: date, symbols: list[str]) -> list[dict[str, Any]]:
    phase1 = _prepare_phase1(_read_csv(_phase1_path(project_root, trade_date)), filter_rate=0.2)
    phase2 = _prepare_phase2(_read_csv(_phase2_path(project_root, trade_date)), filter_rate=0.2)
    phase4 = _prepare_phase4(_read_csv(_phase4_path(project_root, trade_date)), project_root=project_root, trade_date=trade_date)
    phase5 = _prepare_phase5(_read_csv(_phase5_path(project_root)), trade_date=trade_date)
    phase8 = _prepare_phase8(_read_csv(_phase8_path(project_root, trade_date)))
    phase7 = _prepare_phase7(_read_csv(_phase7_path(project_root, trade_date)))
    patterns = _prepare_patterns(_read_csv(_patterns_path(project_root, trade_date)))
    macd = _prepare_symbol_lookup(_read_csv(_macd_path(project_root, trade_date)))
    atr = _prepare_atr(_read_csv(_atr_path(project_root, trade_date)))
    daily_returns = _prepare_daily_returns(_load_daily_returns(project_root=project_root, trade_date=trade_date))
    sectors = _prepare_symbol_lookup(build_sector_display_frame(project_root=project_root))

    rows = [
        _build_output_row(
            symbol=symbol,
            trade_date=trade_date,
            phase1=phase1,
            phase2=phase2,
            phase4=phase4,
            phase5=phase5,
            phase8=phase8,
            phase7=phase7,
            patterns=patterns,
            macd=macd,
            atr=atr,
            daily_returns=daily_returns,
            sectors=sectors,
        )
        for symbol in symbols
    ]

    return rows


def _resolve_workbook_path(project_root: Path, workbook_path: Path | None) -> Path:
    if workbook_path is None:
        return project_root / DEFAULT_TRACK_STOCK_FILENAME
    if workbook_path.is_absolute():
        return workbook_path
    return project_root / workbook_path


def _load_or_create_workbook(path: Path):
    if path.exists():
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
        workbook.active.title = TRACK_INPUT_SHEET
    if TRACK_INPUT_SHEET not in workbook.sheetnames:
        sheet = workbook.create_sheet(TRACK_INPUT_SHEET, 0)
        sheet.append(["股票代码"])
    input_sheet = workbook[TRACK_INPUT_SHEET]
    if input_sheet.max_row == 1 and input_sheet.max_column == 1 and input_sheet["A1"].value is None:
        input_sheet["A1"] = "股票代码"
    _format_input_sheet(input_sheet)
    return workbook


def _read_tracked_symbols(workbook) -> list[str]:
    sheet = workbook[TRACK_INPUT_SHEET]
    header_values = [str(cell.value or "").strip().lower() for cell in sheet[1]]
    symbol_column = 1
    has_symbol_header = False
    for index, value in enumerate(header_values, start=1):
        if value in {"symbol", "code", "股票代码", "代码", "证券代码"}:
            symbol_column = index
            has_symbol_header = True
            break

    start_row = 2 if has_symbol_header else 1
    symbols: list[str] = []
    seen: set[str] = set()
    for row in range(start_row, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=symbol_column)
        cell.number_format = TEXT_NUMBER_FORMAT
        symbol = _normalize_symbol(cell.value)
        if not symbol or symbol in seen:
            continue
        cell.value = symbol
        symbols.append(symbol)
        seen.add(symbol)
    return symbols


def _format_input_sheet(sheet) -> None:
    _trim_empty_tail_rows(sheet)
    header = str(sheet["A1"].value or "").strip().lower()
    if not header or header in {"symbol", "code", "股票代码", "代码", "证券代码"}:
        sheet["A1"] = "股票代码"
    if str(sheet["B1"].value or "").strip().lower() in {"name", "名称", "股票名称"} and _column_is_empty(sheet, 2, start_row=2):
        sheet["B1"] = None
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["A"].number_format = TEXT_NUMBER_FORMAT
    for row in range(1, sheet.max_row + 1):
        sheet.cell(row=row, column=1).number_format = TEXT_NUMBER_FORMAT


def _trim_empty_tail_rows(sheet) -> None:
    last_non_empty = 1
    for row in range(sheet.max_row, 0, -1):
        if any(str(cell.value or "").strip() for cell in sheet[row]):
            last_non_empty = row
            break
    if sheet.max_row > last_non_empty:
        sheet.delete_rows(last_non_empty + 1, sheet.max_row - last_non_empty)


def _column_is_empty(sheet, column: int, *, start_row: int) -> bool:
    for row in range(start_row, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=column).value or "").strip():
            return False
    return True


def _read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path)


def _phase1_path(project_root: Path, trade_date: date) -> Path:
    return project_root / "reports" / "full_market_model" / f"tail_risk_predictions_{trade_date.isoformat()}.csv"


def _phase2_path(project_root: Path, trade_date: date) -> Path:
    return project_root / "reports" / "full_market_model" / f"barrier_risk_predictions_{trade_date.isoformat()}.csv"


def _phase4_path(project_root: Path, trade_date: date) -> Path:
    return project_root / "reports" / "full_market_model" / f"alpha158_qlib_return_predictions_{trade_date.isoformat()}.csv"


def _phase5_path(project_root: Path) -> Path:
    return project_root / "reports" / "full_market_model" / "mcd_crash_annual_measures.csv"


def _phase8_path(project_root: Path, trade_date: date) -> Path:
    return project_root / "reports" / "full_market_model" / f"limit_up_3d_opportunity_predictions_{trade_date.isoformat()}.csv"


def _phase7_path(project_root: Path, trade_date: date) -> Path:
    return project_root / "reports" / "full_market_model" / f"trade_day_gate_prediction_{trade_date.isoformat()}.csv"


def _patterns_path(project_root: Path, trade_date: date) -> Path:
    return project_root / "reports" / "patterns" / f"patterns_all_{trade_date.isoformat()}.csv"


def _macd_path(project_root: Path, trade_date: date) -> Path:
    return project_root / "reports" / "macd" / f"macd_{trade_date.isoformat()}.csv"


def _atr_path(project_root: Path, trade_date: date) -> Path:
    return project_root / "reports" / "atr" / f"atr_{trade_date.isoformat()}.csv"


def _load_daily_returns(*, project_root: Path, trade_date: date) -> pd.DataFrame:
    frame = read_full_market_daily_returns(project_root=project_root, trade_date=trade_date)
    if not frame.empty:
        return frame
    write_full_market_daily_returns(project_root=project_root, trade_date=trade_date)
    return read_full_market_daily_returns(project_root=project_root, trade_date=trade_date)


def _prepare_daily_returns(frame: pd.DataFrame) -> dict[str, dict[str, Any]]:
    if frame.empty or "symbol" not in frame.columns:
        return {}
    data = frame.copy()
    data["symbol"] = data["symbol"].map(_normalize_symbol)
    keep = [
        "symbol",
        "daily_return_pct",
        "daily_close",
        "daily_prev_close",
        "daily_volume",
        "daily_amount",
    ]
    return _records_by_symbol(data.dropna(subset=["symbol"]), keep)


def _prepare_phase1(frame: pd.DataFrame, *, filter_rate: float) -> dict[str, dict[str, Any]]:
    return _prepare_risk_scores(
        frame,
        source_score="risk_score",
        output_score="phase1_risk_score",
        rank_column="phase1_risk_rank",
        percentile_column="phase1_risk_percentile",
        excluded_column="phase1_excluded_top20",
        filter_rate=filter_rate,
    )


def _prepare_phase2(frame: pd.DataFrame, *, filter_rate: float) -> dict[str, dict[str, Any]]:
    prepared = _prepare_risk_scores(
        frame,
        source_score="barrier_risk_score",
        output_score="phase2_barrier_risk_score",
        rank_column="phase2_risk_rank",
        percentile_column="phase2_risk_percentile",
        excluded_column="phase2_excluded_top20",
        filter_rate=filter_rate,
    )
    if frame.empty or "symbol" not in frame.columns:
        return prepared
    extras = frame.copy()
    extras["symbol"] = extras["symbol"].map(_normalize_symbol)
    extras = extras.drop_duplicates("symbol", keep="first")
    for _, row in extras.iterrows():
        symbol = str(row.get("symbol", ""))
        if not symbol:
            continue
        prepared.setdefault(symbol, {})
        if "is_cusum_event" in row.index:
            prepared[symbol]["phase2_is_cusum_event"] = _normalize_value(row.get("is_cusum_event"))
    return prepared


def _prepare_risk_scores(
    frame: pd.DataFrame,
    *,
    source_score: str,
    output_score: str,
    rank_column: str,
    percentile_column: str,
    excluded_column: str,
    filter_rate: float,
) -> dict[str, dict[str, Any]]:
    if frame.empty or "symbol" not in frame.columns or source_score not in frame.columns:
        return {}
    data = frame.copy()
    data["symbol"] = data["symbol"].map(_normalize_symbol)
    data[output_score] = pd.to_numeric(data[source_score], errors="coerce")
    data = data.dropna(subset=["symbol", output_score]).sort_values([output_score, "symbol"], ascending=[False, True])
    data = data.drop_duplicates("symbol", keep="first").reset_index(drop=True)
    if data.empty:
        return {}
    data[rank_column] = data.index + 1
    data[percentile_column] = data[output_score].rank(pct=True, method="max")
    score_column = f"{rank_column.rsplit('_risk_rank', 1)[0]}_score_100"
    data[score_column] = score_series_100(data[output_score], higher_is_better=False)
    removed_rows = max(1, int(math.ceil(len(data) * filter_rate)))
    data[excluded_column] = False
    data.loc[: removed_rows - 1, excluded_column] = True
    keep = ["symbol", "name", score_column, output_score, rank_column, percentile_column, excluded_column]
    return _records_by_symbol(data, keep)


def _prepare_phase4(frame: pd.DataFrame, *, project_root: Path, trade_date: date) -> dict[str, dict[str, Any]]:
    if frame.empty or "symbol" not in frame.columns or "return_score" not in frame.columns:
        return {}
    data = frame.copy()
    data["symbol"] = data["symbol"].map(_normalize_symbol)
    data["phase4_return_score"] = pd.to_numeric(data["return_score"], errors="coerce")
    data = data.dropna(subset=["symbol", "phase4_return_score"]).sort_values(["phase4_return_score", "symbol"], ascending=[False, True])
    data = data.drop_duplicates("symbol", keep="first").reset_index(drop=True)
    if data.empty:
        return {}
    data["phase4_rank"] = data.index + 1
    data["phase4_score_percentile"] = data["phase4_return_score"].rank(pct=True, method="max")
    data["phase4_score_100"] = score_series_100(data["phase4_return_score"], higher_is_better=True)
    data = merge_phase4_rolling_frame(data, project_root=project_root, trade_date=trade_date)
    keep = [
        "symbol",
        "name",
        "phase4_score_100",
        "phase4_5d_mean",
        "phase4_5d_std",
        "phase4_return_score",
        "phase4_rank",
        "phase4_score_percentile",
    ]
    return _records_by_symbol(data, keep)


def _prepare_phase5(frame: pd.DataFrame, *, trade_date: date) -> dict[str, dict[str, Any]]:
    if frame.empty or "symbol" not in frame.columns:
        return {}
    data = frame.copy()
    data["symbol"] = data["symbol"].map(_normalize_symbol)
    data["year"] = pd.to_numeric(data.get("year"), errors="coerce")
    data = data.dropna(subset=["symbol", "year"])
    if data.empty:
        return {}
    eligible = data[data["year"].astype(int).le(trade_date.year)].copy()
    if eligible.empty:
        eligible = data.copy()
    eligible = eligible.sort_values(["symbol", "year"]).drop_duplicates("symbol", keep="last")
    eligible = eligible.rename(
        columns={
            "NEGOUTLIER": "phase5_NEGOUTLIER",
            "CRASH": "phase5_CRASH",
            "CRASH_count": "phase5_CRASH_count",
            "NCSKEW": "phase5_NCSKEW",
            "DUVOL": "phase5_DUVOL",
            "RET": "phase5_RET",
            "SIGMA": "phase5_SIGMA",
            "MINRET": "phase5_MINRET",
            "year": "phase5_year",
        }
    )
    eligible = add_phase5_score_100(eligible)
    keep = [
        "symbol",
        "phase5_score_100",
        "phase5_NEGOUTLIER",
        "phase5_CRASH",
        "phase5_CRASH_count",
        "phase5_NCSKEW",
        "phase5_DUVOL",
        "phase5_RET",
        "phase5_SIGMA",
        "phase5_MINRET",
        "phase5_year",
    ]
    records: dict[str, dict[str, Any]] = {}
    for symbol, row in _records_by_symbol(eligible, keep).items():
        records[symbol] = row
    return records


def _prepare_phase8(frame: pd.DataFrame) -> dict[str, dict[str, Any]]:
    if frame.empty or "symbol" not in frame.columns:
        return {}
    data = frame.copy()
    data["symbol"] = data["symbol"].map(_normalize_symbol)
    if "phase8_score_100" not in data.columns:
        source_column = "phase8_raw_score" if "phase8_raw_score" in data.columns else "limit_up_opportunity_score"
        if source_column not in data.columns:
            return {}
        data["phase8_score_100"] = score_series_100(pd.to_numeric(data[source_column], errors="coerce"), higher_is_better=True)
    if "phase8_raw_score" not in data.columns and "limit_up_opportunity_score" in data.columns:
        data["phase8_raw_score"] = pd.to_numeric(data["limit_up_opportunity_score"], errors="coerce")
    data["phase8_score_100"] = pd.to_numeric(data["phase8_score_100"], errors="coerce")
    data = data.dropna(subset=["symbol", "phase8_score_100"]).sort_values(["phase8_score_100", "symbol"], ascending=[False, True])
    data = data.drop_duplicates("symbol", keep="first").reset_index(drop=True)
    if data.empty:
        return {}
    if "phase8_rank" not in data.columns:
        data["phase8_rank"] = data.index + 1
    keep = ["symbol", "phase8_score_100", "phase8_raw_score", "phase8_rank", "today_limit_up_excluded"]
    return _records_by_symbol(data, keep)


def _prepare_phase7(frame: pd.DataFrame) -> dict[str, Any]:
    if frame.empty:
        return {}
    row = frame.iloc[0]
    permission = row.get("trade_permission")
    return {
        "phase7_score_100": _normalize_value(phase7_score_100(permission)),
        "phase7_trade_permission": _normalize_value(row.get("trade_permission")),
        "phase7_buy_day_risk_score": _normalize_value(row.get("buy_day_risk_score")),
        "phase7_selected_threshold": _normalize_value(row.get("selected_threshold")),
        "phase7_suggested_action": _normalize_value(row.get("suggested_action")),
        "phase7_reason": _normalize_value(row.get("reason")),
    }


def _prepare_patterns(frame: pd.DataFrame) -> dict[str, dict[str, Any]]:
    if frame.empty or "symbol" not in frame.columns or "pattern_id" not in frame.columns:
        return {}
    data = frame.copy()
    data["symbol"] = data["symbol"].map(_normalize_symbol)
    data["pattern_id"] = data["pattern_id"].astype(str)
    records: dict[str, dict[str, Any]] = {}
    for symbol, group in data.dropna(subset=["symbol"]).groupby("symbol", sort=False):
        ids = sorted({str(value).strip() for value in group["pattern_id"].dropna().tolist() if str(value).strip()})
        item = {
            "pattern_match": "是" if ids else "",
            "pattern_ids": ",".join(ids),
        }
        id_set = set(ids)
        for pattern_id in range(1, 7):
            item[f"pattern{pattern_id}"] = "是" if str(pattern_id) in id_set else ""
        records[str(symbol)] = item
    return records


def _prepare_symbol_lookup(frame: pd.DataFrame) -> dict[str, dict[str, Any]]:
    if frame.empty or "symbol" not in frame.columns:
        return {}
    data = frame.copy()
    data["symbol"] = data["symbol"].map(_normalize_symbol)
    return _records_by_symbol(data.dropna(subset=["symbol"]), list(data.columns))


def _prepare_atr(frame: pd.DataFrame) -> dict[str, dict[str, Any]]:
    if frame.empty:
        return {}
    rename_map = {
        "代码": "symbol",
        "名称": "name",
        "交易日期": "trade_date",
        "收盘价": "close",
        "ATR14": "atr_14",
        "ATR%": "atr_pct_14",
        "1ATR止损参考": "atr_stop_loss_1x",
        "2ATR止损参考": "atr_stop_loss_2x",
        "2ATR止盈参考": "atr_take_profit_2x",
        "3ATR止盈参考": "atr_take_profit_3x",
        "波动分层": "atr_volatility_regime",
    }
    data = frame.rename(columns={key: value for key, value in rename_map.items() if key in frame.columns}).copy()
    return _prepare_symbol_lookup(data)


def _records_by_symbol(frame: pd.DataFrame, columns: list[str]) -> dict[str, dict[str, Any]]:
    available = [column for column in columns if column in frame.columns]
    if "symbol" not in available:
        return {}
    data = frame.loc[:, available].drop_duplicates("symbol", keep="first")
    records: dict[str, dict[str, Any]] = {}
    for _, row in data.iterrows():
        symbol = _normalize_symbol(row.get("symbol"))
        if not symbol:
            continue
        records[symbol] = {column: _normalize_value(row.get(column)) for column in available if column != "symbol"}
    return records


def _build_output_row(
    *,
    symbol: str,
    trade_date: date,
    phase1: dict[str, dict[str, Any]],
    phase2: dict[str, dict[str, Any]],
    phase4: dict[str, dict[str, Any]],
    phase5: dict[str, dict[str, Any]],
    phase8: dict[str, dict[str, Any]],
    phase7: dict[str, Any],
    patterns: dict[str, dict[str, Any]],
    macd: dict[str, dict[str, Any]],
    atr: dict[str, dict[str, Any]],
    daily_returns: dict[str, dict[str, Any]],
    sectors: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    row: dict[str, Any] = {column: "" for column in TRACK_STOCK_COLUMNS}
    row["trade_date"] = trade_date.isoformat()
    row["symbol"] = symbol
    for source in (daily_returns, phase1, phase2, phase4, phase5, phase8, sectors, patterns, macd, atr):
        row.update(source.get(symbol, {}))
    row.update(phase7)
    position_pct = recommended_position_percent_from_mapping(row)
    if position_pct is not None:
        row["recommended_position_pct"] = position_pct
    return {column: row.get(column, "") for column in TRACK_STOCK_COLUMNS}


def _build_intraday_rows(
    *,
    project_root: Path,
    trade_date: date,
    symbols: list[str],
    intraday_frame: pd.DataFrame | None,
) -> list[dict[str, Any]]:
    frame = intraday_frame.copy() if intraday_frame is not None else _read_intraday_track_source(project_root, trade_date)
    prepared = _prepare_intraday_track_frame(frame, trade_date=trade_date)
    records = _records_by_symbol(prepared, INTRADAY_TRACK_STOCK_COLUMNS)
    rows: list[dict[str, Any]] = []
    for symbol in symbols:
        row: dict[str, Any] = {column: "" for column in INTRADAY_TRACK_STOCK_COLUMNS}
        row["intraday_trade_date"] = trade_date.isoformat()
        row["symbol"] = symbol
        row.update(records.get(symbol, {}))
        rows.append({column: row.get(column, "") for column in INTRADAY_TRACK_STOCK_COLUMNS})
    return rows


def _read_intraday_track_source(project_root: Path, trade_date: date) -> pd.DataFrame:
    report_dir = project_root / "reports" / "intraday_screening"
    preferred = report_dir / f"intraday_track_stock_{trade_date.isoformat()}.csv"
    if preferred.exists():
        return _read_csv(preferred)
    fallback = report_dir / f"intraday_pool_screening_{trade_date.isoformat()}.csv"
    if fallback.exists():
        return _read_csv(fallback)
    return pd.DataFrame()


def _prepare_intraday_track_frame(frame: pd.DataFrame, *, trade_date: date) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame(columns=INTRADAY_TRACK_STOCK_COLUMNS)
    data = frame.rename(columns={key: value for key, value in INTRADAY_CHINESE_TO_INTERNAL.items() if key in frame.columns}).copy()
    if "symbol" not in data.columns:
        return pd.DataFrame(columns=INTRADAY_TRACK_STOCK_COLUMNS)
    data["symbol"] = data["symbol"].map(_normalize_symbol)
    data = data[data["symbol"].astype(str).str.len().eq(6)].copy()
    if data.empty:
        return pd.DataFrame(columns=INTRADAY_TRACK_STOCK_COLUMNS)
    if "intraday_trade_date" not in data.columns:
        data["intraday_trade_date"] = trade_date.isoformat()
    if "intraday_selection_source" not in data.columns:
        data["intraday_selection_source"] = data.get("prev_source", "")
    if "intraday_pool_score" not in data.columns and "centered_risk_score" in data.columns:
        data["intraday_pool_score"] = data["centered_risk_score"]
    if "recommended_position_pct" not in data.columns:
        data[RECOMMENDED_POSITION_PERCENT_FIELD] = data.get(RECOMMENDED_POSITION_PERCENT_FIELD, "")
        data["recommended_position_pct"] = data[RECOMMENDED_POSITION_PERCENT_FIELD]
    for column in INTRADAY_TRACK_STOCK_COLUMNS:
        if column not in data.columns:
            data[column] = ""
    return data.loc[:, INTRADAY_TRACK_STOCK_COLUMNS].drop_duplicates("symbol", keep="first")


def _write_output_sheet(
    workbook,
    rows: list[dict[str, Any]],
    *,
    sheet_name: str,
    columns: list[str],
    headers: dict[str, str],
) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    index = 1 if sheet_name == TRACK_DAILY_OUTPUT_SHEET else 2 if sheet_name == TRACK_INTRADAY_OUTPUT_SHEET else None
    sheet = workbook.create_sheet(sheet_name, index=index)
    sheet.append([headers.get(column, column) for column in columns])
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        width = max(10, min(max(len(str(column)) + 2, 12), 28))
        sheet.column_dimensions[get_column_letter(index)].width = width
        if column == "symbol":
            sheet.column_dimensions[get_column_letter(index)].number_format = TEXT_NUMBER_FORMAT
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=1, max_row=sheet.max_row):
                for item in cell:
                    item.number_format = TEXT_NUMBER_FORMAT


def _ensure_sheet_order(workbook) -> None:
    if TRACK_DAILY_OUTPUT_SHEET not in workbook.sheetnames:
        sheet = workbook.create_sheet(TRACK_DAILY_OUTPUT_SHEET, index=1)
        sheet.append([TRACK_STOCK_HEADERS_ZH.get(column, column) for column in TRACK_STOCK_COLUMNS])
        sheet.freeze_panes = "A2"
    if TRACK_INTRADAY_OUTPUT_SHEET not in workbook.sheetnames:
        sheet = workbook.create_sheet(TRACK_INTRADAY_OUTPUT_SHEET, index=2)
        sheet.append([INTRADAY_TRACK_STOCK_HEADERS_ZH.get(column, column) for column in INTRADAY_TRACK_STOCK_COLUMNS])
        sheet.freeze_panes = "A2"
    desired = [TRACK_INPUT_SHEET, TRACK_DAILY_OUTPUT_SHEET, TRACK_INTRADAY_OUTPUT_SHEET]
    ordered = [workbook[name] for name in desired if name in workbook.sheetnames]
    ordered.extend([sheet for sheet in workbook.worksheets if sheet.title not in desired])
    workbook._sheets = ordered


def _normalize_symbol(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    if text.startswith('="') and text.endswith('"'):
        text = text[2:-1]
    if text.startswith("="):
        text = text.lstrip("=").strip().strip('"')
    text = text.replace(".0", "") if text.endswith(".0") else text
    lower = text.lower()
    if lower.startswith(("sh", "sz", "bj")):
        text = text[2:]
    digits = "".join(char for char in text if char.isdigit())
    if not digits:
        return ""
    return digits[-6:].zfill(6)


def _normalize_value(value: object) -> Any:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return round(value, 6)
    return value
